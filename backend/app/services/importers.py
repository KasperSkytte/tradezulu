"""Importers for people who cannot (or will not) run the Expert Advisor.

Two formats are understood:

* the HTML trade report MetaTrader 5 writes from *History -> Report*, whose
  "Positions" table already is trade-level data;
* a generic CSV, either trade-level or deal-level, with flexible headers.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from html.parser import HTMLParser
from typing import Any

NBSP = " "
NNBSP = " "


class _TableRowParser(HTMLParser):
    """Collect every table row as a list of cell strings."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)


def _clean_number(text: str) -> float:
    text = (text or "").replace(NBSP, "").replace(NNBSP, "").replace(" ", "").replace(",", "")
    text = text.replace("−", "-")  # MT5 sometimes emits a unicode minus
    if not text or text in {"-", "--"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


DATE_PATTERNS = (
    "%Y.%m.%d %H:%M:%S",
    "%Y.%m.%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%d.%m.%Y %H:%M:%S",
    "%m/%d/%Y %H:%M:%S",
)


def _parse_dt(text: str) -> datetime | None:
    text = (text or "").replace(NBSP, " ").strip()
    if not text:
        return None
    for pattern in DATE_PATTERNS:
        try:
            return datetime.strptime(text, pattern).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def position_from_row(cells: list[str]) -> dict[str, Any] | None:
    """One row of a "Positions" table, or None if it is not one.

    Mapped by shape rather than by column number. The same report is written
    with different columns depending on where it came from: the HTML export
    carries an unnamed comment cell after Type that the visible header does not
    mention, so its rows are 14 wide against a 13-column header, while the
    spreadsheet export of the very same history has no such cell. Fixed indices
    read one file correctly and silently shift every field of the other --
    volume into price, price into the stop, the close price into commission.

    Two things are reliable in every variant: the row opens with a timestamp,
    and the close time is the next timestamp along. Everything between them is
    volume, price, stop and target in that order, preceded by a comment if the
    export has one; everything after is close price, commission, swap, profit.
    """
    if len(cells) < 13:
        return None

    opened_at = _parse_dt(cells[0])
    if opened_at is None:
        return None

    direction = cells[3].strip().lower()
    if direction not in ("buy", "sell"):
        return None

    closed_index = next(
        (i for i in range(4, len(cells)) if _parse_dt(cells[i]) is not None), None
    )
    if closed_index is None:
        return None

    # Take the last four, so an optional comment ahead of them is skipped.
    middle = cells[4:closed_index]
    if len(middle) < 4:
        return None
    volume, entry_price, stop, target = middle[-4:]

    tail = cells[closed_index + 1 :]
    if len(tail) < 4:
        return None
    exit_price, commission, swap, gross_profit = tail[:4]

    try:
        position_id = int(_clean_number(cells[1]))
    except (TypeError, ValueError):
        return None

    return {
        "position_id": position_id,
        "symbol": cells[2].strip().upper(),
        "direction": "long" if direction == "buy" else "short",
        "volume": _clean_number(volume),
        "entry_price": _clean_number(entry_price),
        "initial_stop": _clean_number(stop) or None,
        "initial_target": _clean_number(target) or None,
        "opened_at": opened_at,
        "closed_at": _parse_dt(cells[closed_index]),
        "exit_price": _clean_number(exit_price),
        "commission": _clean_number(commission),
        "swap": _clean_number(swap),
        "gross_profit": _clean_number(gross_profit),
    }


#: Words that describe the account rather than name it: they appear in the same
#: parenthesised list and must not be mistaken for the holder.
_ACCOUNT_KEYWORDS = {"demo", "real", "contest", "hedge", "hedging", "netting", "preliminary"}


def _account_from_rows(rows: list[list[str]], account: dict[str, Any]) -> None:
    """Read the report's header block into ``account``, in place.

    ``Account: 25702871 (USD, VantageMarkets-Demo, demo, Hedge)`` -- the
    parenthesised parts are currency, server, kind and netting mode, and they
    are identified by what they look like rather than by position, because
    which of them a build emits varies. The holder's name is on its own row and
    was previously read out of the parentheses, which is how an import ended up
    filed under a person called "USD".
    """
    for cells in rows:
        # Empties dropped here, unlike in the trade rows: the spreadsheet pads
        # this block with blank cells, so "Name:" and its value are not
        # necessarily neighbours by index.
        text = [c.replace(NBSP, " ").strip() for c in cells if c.replace(NBSP, " ").strip()]
        joined = " ".join(text)
        if not joined:
            continue

        match = re.search(r"Account:?\s*([0-9]{4,})", joined)
        if match and account["login"] == "0":
            account["login"] = match.group(1)
            inside = re.search(r"\(([^)]*)\)", joined)
            for part in (p.strip() for p in (inside.group(1) if inside else "").split(",")):
                if not part:
                    continue
                if re.fullmatch(r"[A-Z]{3}", part):
                    account["currency"] = part
                elif "-" in part and not account["server"]:
                    account["server"] = part[:120]
                elif part.lower() in _ACCOUNT_KEYWORDS:
                    continue
                elif not account["name"]:
                    # Some builds lead with the holder's name, others do not
                    # carry it here at all. An explicit "Name:" row wins.
                    account["name"] = part[:120]

        # Authoritative when present, so it replaces anything guessed above.
        if len(text) >= 2 and text[0].rstrip(":").lower() == "name":
            account["name"] = text[1][:120]
        if len(text) >= 2 and text[0].rstrip(":").lower() in {"company", "broker"}:
            account.setdefault("broker", "")
            if not account["broker"]:
                account["broker"] = text[1][:120]


def parse_report_rows(rows: list[list[str]]) -> dict[str, Any]:
    """Account details and closed positions from a MetaTrader report's rows.

    Shared by the HTML and spreadsheet exports, which differ only in how the
    rows are read off the page.
    """
    account: dict[str, Any] = {
        "login": "0", "server": "", "name": "", "currency": "USD", "broker": "",
    }
    _account_from_rows(rows, account)

    positions: list[dict[str, Any]] = []
    section = ""
    for row in rows:
        cells = [c.replace(NBSP, " ").strip() for c in row]
        named = [c for c in cells if c]
        if not named:
            continue

        if len(named) == 1:
            # "Open Positions" is a different table with a different shape, and
            # holds trades that have not finished. Only closed ones are history.
            section = named[0].lower().strip(":")
            continue

        if section != "positions":
            continue

        position = position_from_row(cells)
        if position is not None:
            positions.append(position)

    return {"account": account, "positions": positions}


def parse_mt5_html_report(content: str) -> dict[str, Any]:
    """Extract account info and positions from an MT5 HTML report."""
    parser = _TableRowParser()
    parser.feed(content)
    return parse_report_rows(parser.rows)


def parse_mt5_xlsx_report(data: bytes) -> dict[str, Any]:
    """The same report saved as a spreadsheet.

    MetaTrader offers XLSX beside HTML in the same menu, and it is the one
    people reach for. It was not accepted at all before: the upload was decoded
    as text, which for a zip container produces nothing that looks like a
    trade.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ValueError(
            "Reading .xlsx reports needs the openpyxl package, which is missing "
            "from this build. Save the report as HTML instead."
        ) from exc

    import warnings

    with warnings.catch_warnings():
        # MetaTrader writes no default style; openpyxl says so on every load.
        warnings.simplefilter("ignore")
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        rows: list[list[str]] = []
        for sheet in workbook.worksheets:
            for raw in sheet.iter_rows(values_only=True):
                rows.append(["" if value is None else str(value).strip() for value in raw])
        workbook.close()

    return parse_report_rows(rows)


# --- CSV --------------------------------------------------------------------

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "position_id": ("position", "position_id", "positionid", "ticket", "id", "trade_id"),
    "symbol": ("symbol", "instrument", "market", "pair"),
    "direction": ("direction", "type", "side", "buy/sell", "long/short"),
    "volume": ("volume", "size", "lots", "quantity", "qty"),
    "entry_price": ("entry", "entry_price", "openprice", "open_price", "price_open", "price"),
    "exit_price": ("exit", "exit_price", "closeprice", "close_price", "price_close"),
    "opened_at": ("opentime", "open_time", "open time", "time_open", "opened_at", "entrytime", "date"),
    "closed_at": ("closetime", "close_time", "close time", "time_close", "closed_at", "exittime"),
    "initial_stop": ("sl", "s/l", "stop", "stoploss", "stop_loss", "initial_stop"),
    "initial_target": ("tp", "t/p", "target", "takeprofit", "take_profit", "initial_target"),
    "gross_profit": ("profit", "grossprofit", "gross_profit", "pnl", "p/l", "net"),
    "commission": ("commission", "commissions", "fee", "fees"),
    "swap": ("swap", "rollover", "interest"),
    "notes": ("notes", "comment", "note"),
    "setup": ("setup", "strategy"),
    "tags": ("tags", "tag"),
}


def _normalise_header(name: str) -> str:
    return re.sub(r"[^a-z0-9/_ ]", "", (name or "").strip().lower())


def _build_column_map(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalised = [_normalise_header(h) for h in headers]
    for field, aliases in COLUMN_ALIASES.items():
        for index, header in enumerate(normalised):
            if header in aliases and field not in mapping:
                mapping[field] = index
                break
    return mapping


def parse_trades_csv(content: str) -> list[dict[str, Any]]:
    """Parse a trade-level CSV into position dicts."""
    text = content.lstrip("﻿")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)

    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return []

    mapping = _build_column_map(rows[0])
    if "symbol" not in mapping:
        raise ValueError("Could not find a symbol column in the CSV")

    out: list[dict[str, Any]] = []
    for index, row in enumerate(rows[1:], start=1):
        def cell(field: str, default: str = "", row: list[str] = row) -> str:
            position = mapping.get(field)
            if position is None or position >= len(row):
                return default
            return row[position].strip()

        symbol = cell("symbol").upper()
        if not symbol:
            continue

        direction_raw = cell("direction").lower()
        if direction_raw.startswith(("b", "l")):
            direction = "long"
        elif direction_raw.startswith(("s",)):
            direction = "short"
        else:
            direction = "long"

        opened_at = _parse_dt(cell("opened_at"))
        if opened_at is None:
            continue

        out.append(
            {
                "position_id": int(_clean_number(cell("position_id"))) or -index,
                "symbol": symbol,
                "direction": direction,
                "volume": _clean_number(cell("volume")) or 1.0,
                "entry_price": _clean_number(cell("entry_price")),
                "exit_price": _clean_number(cell("exit_price")) or None,
                "opened_at": opened_at,
                "closed_at": _parse_dt(cell("closed_at")),
                "initial_stop": _clean_number(cell("initial_stop")) or None,
                "initial_target": _clean_number(cell("initial_target")) or None,
                "gross_profit": _clean_number(cell("gross_profit")),
                "commission": _clean_number(cell("commission")),
                "swap": _clean_number(cell("swap")),
                "notes": cell("notes"),
                "setup": cell("setup"),
                "tags": [t.strip() for t in cell("tags").replace("|", ",").split(",") if t.strip()],
            }
        )
    return out
